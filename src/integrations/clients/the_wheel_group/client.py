"""
Client for The Wheel Group's ``US Wheel Data Mastersheet.xlsx``.

TWG publishes one workbook holding their whole US catalog. Only the ``US Data Mastersheet``
worksheet is read: one row per SKU across TWG's house brands (Touren, Mayhem, ION Alloy, Cali
Off-Road, Ridler, Dirty Life, Kraze, American Truxx, Mazzi, TuffStuff, ION Trailer), carrying the
wheel specs, images, marketing copy, MSRP and MAP. The workbook's other sheets are deliberately
ignored -- ``Discontinued`` is a retired-SKU archive, ``WINTER WHEEL PROMO`` a seasonal price list,
and ``Version Changes`` a changelog.

There is **no stock and no dealer cost** in this workbook: it is a catalog + list-price sheet.
Availability and per-dealer cost only arrive once TWG pushes a real feed to our relay.

Two transports, same workbook parser:

  * ``public_share`` (default today) -- TWG's public Dropbox folder share. Requesting it with
    ``dl=1`` returns the whole folder as a zip; the mastersheet is extracted from it in memory.
    The share URL lives in ``settings.THE_WHEEL_GROUP_PUBLIC_SHARE_URL`` and can be overridden per
    connection via ``credentials['feed']['public_share_url']``.
  * ``sftp`` -- our own relay (the same ``5.161.121.143:22`` / ``uploads`` endpoint Meyer and
    A-Tech use). TWG is a relay-provisioned integration: the dealer enters nothing, and their
    connection carries the relay account we generated for them (``ftp_user`` / ``ftp_password``,
    see ``constants.PROVIDERS_CATALOG`` -> ``relay_credential_fields``). Nothing is pushed there
    yet, which is why ``THE_WHEEL_GROUP_FORCE_PUBLIC_SHARE`` defaults to True -- flip it off once
    TWG starts delivering, and every connection switches to its own drop with no code change.

Both transports accept either a bare ``.xlsx`` or a ``.zip`` containing one, so the Dropbox zip
and a future direct workbook drop go through the same path.
"""
import io
import logging
import os
import re
import tempfile
import typing
import urllib.error
import urllib.request
import zipfile

import openpyxl
from django.conf import settings

from src.integrations.clients.the_wheel_group import exceptions

logger = logging.getLogger(__name__)

_LOG_PREFIX = "[THE-WHEEL-GROUP-CLIENT]"

# TWG's public Dropbox folder share. dl=1 makes Dropbox serve the folder as a zip rather than the
# HTML preview page. Kept as the default here as well as in settings so the client still works in
# a shell/test context with no settings override.
DEFAULT_PUBLIC_SHARE_URL = (
    "https://www.dropbox.com/scl/fo/dwls6ye5utcah34ptw61n/"
    "AKgB7W8qs2E8h_C9uzxaKqs?rlkey=nud3j0hudfaakd3m9enf8ag35&dl=1"
)
DEFAULT_RELAY_HOST = "5.161.121.143"
DEFAULT_RELAY_PORT = 22
DEFAULT_RELAY_DIRECTORY = "uploads"

SOURCE_MODE_PUBLIC_SHARE = "public_share"
SOURCE_MODE_SFTP = "sftp"

# The only worksheet we ingest -- see the module docstring for why the others are skipped.
DATA_SHEET_NAME = "US Data Mastersheet"

# Workbook filename on the relay. TWG names the file "US Wheel Data Mastersheet.xlsx" in the
# Dropbox share; the relay drop is not live yet, so anything that reads as a mastersheet is
# accepted and only the newest one is used.
WORKBOOK_FILENAME_PATTERN = re.compile(r"master.*\.(xlsx|zip)$", re.IGNORECASE)
_WORKBOOK_SUFFIXES = (".xlsx", ".zip")

_REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; AfterMarketScout/1.0; +https://aftermarketscout.com)",
    "Accept": "*/*",
}
_DOWNLOAD_TIMEOUT_SECONDS = 300
_HEAD_TIMEOUT_SECONDS = 60
_SFTP_CONNECT_TIMEOUT_SECONDS = 15

# openpyxl hands back whatever the cell holds, including Excel's own error sentinels for the few
# formula cells TWG ships broken (one #NAME? in OFFSETNUM, two #VALUE! in WHEEL_LIP_SIZE). They
# are not values -- they read as literal "#VALUE!" text in every downstream field if kept.
_EXCEL_ERROR_VALUES = frozenset(
    {"#NAME?", "#VALUE!", "#REF!", "#N/A", "#DIV/0!", "#NULL!", "#NUM!"}
)

_HEADER_KEY_RE = re.compile(r"[^A-Z0-9]")

# Compact header key (letters/digits only, uppercased) -> our field name. Compact keys absorb the
# spacing and punctuation drift between sheets: the mastersheet writes "SHIP WEIGHT(LBS)" and
# "IMAGE1", the Discontinued sheet "SHIP WEIGHT (LBS)" and "IMAGE URL 1". Aliases are listed for
# every label TWG is known to use for the same column so a re-titled export keeps parsing.
_FIELD_BY_HEADER_KEY: typing.Dict[str, str] = {
    "SKU": "sku",
    "BRAND": "brand",
    "AAIACODE": "aaia_code",
    "NAME": "name",
    "STYLENUMBER": "style_number",
    "DESCRIPTION": "description",
    "SHORTDESCRIPTION": "short_description",
    "DIAMETER": "diameter",
    "WHEELWIDTH": "wheel_width",
    "HUB": "hub_bore",
    "PCD1": "bolt_pattern_1",
    "PCD2": "bolt_pattern_2",
    "OFFSET": "offset_class",
    "OFFSETNUM": "offset",
    "ACCESSORY": "accessory",
    "WHEELCAP": "wheel_cap",
    "SCREW": "screw",
    "DUALLYWHEEL": "dually_wheel",
    "WINTERAPPROVED": "winter_approved",
    "COLOR": "color",
    "FINISH": "finish",
    "LOADRATING": "load_rating",
    "UPC": "upc",
    "COUNTRYOFORIGIN": "country_of_origin",
    "DIVISION": "division",
    "GROUPCODE": "group_code",
    "NOTE": "note",
    "COMMENT": "comment",
    "BACKSPACE": "backspace",
    "STRUCTUREWARRANTY": "structure_warranty",
    "FINISHWARRANTY": "finish_warranty",
    "BEADLOCKINSTALLATIONINSTRUCTIONS": "beadlock_instructions_url",
    "TPMSCOMPATIBLE": "tpms_compatible",
    "WHEELLIPSIZE": "wheel_lip_size",
    "LUGNUTOPENCLOSED": "lugnut_open_closed",
    "LUGNUTTYPE1": "lugnut_type_1",
    "LUGNUTTYPE2": "lugnut_type_2",
    "LUGSEATTYPE": "lugseat_type",
    "WIDTH": "box_width",
    "HEIGHT": "box_height",
    "DEPTH": "box_depth",
    "PRODUCTWEIGHTLBS": "product_weight",
    "SHIPWEIGHTLBS": "ship_weight",
    "IMAGE1": "image_1",
    "IMAGE2": "image_2",
    "IMAGE3": "image_3",
    "IMAGE4": "image_4",
    "IMAGEURL1": "image_1",
    "IMAGEURL2": "image_2",
    "IMAGEURL3": "image_3",
    "IMAGEURL4": "image_4",
    "BULLETPOINTS": "bullet_points",
    "SALESDESCRIPTION": "sales_description",
    "NEWMSRPPRICE": "msrp",
    "MSRP": "msrp",
    "NEWMAPPRICE": "map_price",
    "USMAP": "map_price",
    "MAPYN": "map_enforced",
    "MAPYESNO": "map_enforced",
    # Dealer cost. The public mastersheet has no such column -- these are the spellings TWG uses
    # elsewhere in the same workbook ("Jobber Price" on the promo sheet) and the usual industry
    # variants, so a per-dealer feed delivered to the relay prices itself without a code change.
    # Only the per-company pricing path reads this; it never lands on TheWheelGroupPart, which is
    # shared across companies.
    "JOBBERPRICE": "cost",
    "JOBBER": "cost",
    "DEALERPRICE": "cost",
    "DEALERCOST": "cost",
    "YOURPRICE": "cost",
    "WHOLESALEPRICE": "cost",
    "WHOLESALE": "cost",
    "COST": "cost",
}

# The header row is found by scanning for the row that carries these -- the mastersheet opens with
# a free-text price-change banner and a blank spacer row before it.
_REQUIRED_HEADER_KEYS = ("SKU", "BRAND")


def _header_key(value: typing.Any) -> str:
    return _HEADER_KEY_RE.sub("", str(value or "").upper())


def _text(value: typing.Any) -> typing.Optional[str]:
    """Trimmed cell text, or None for blanks and Excel error sentinels."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        # openpyxl reads 20 as 20.0 -- keep whole numbers integral so SKUs, sizes and UPCs don't
        # pick up a ".0" tail.
        text = str(int(value))
    else:
        text = str(value).strip()
    text = text.strip()
    if not text or text.upper() in _EXCEL_ERROR_VALUES:
        return None
    # TWG's UPC column is text-formatted in places and keeps Excel's leading apostrophe.
    return text.lstrip("'").strip() or None


class TheWheelGroupFeedClient(object):
    """
    Downloads TWG's newest mastersheet and returns raw row dicts (numeric coercion and model
    mapping are the service layer's job, matching the Elite Wheel and Vossen clients).
    """

    def __init__(
        self,
        credentials: typing.Optional[typing.Dict] = None,
        local_file_path: typing.Optional[str] = None,
        force_public_share: typing.Optional[bool] = None,
    ):
        creds = credentials or {}
        self.public_share_url = (
            str(creds.get("public_share_url") or "").strip()
            or getattr(settings, "THE_WHEEL_GROUP_PUBLIC_SHARE_URL", None)
            or DEFAULT_PUBLIC_SHARE_URL
        )
        # Relay-provisioned connections store the account we generated as ftp_user/ftp_password
        # (see PROVIDERS_CATALOG relay_credential_fields); sftp_* is accepted too so this client
        # also works with a hand-entered account.
        self.sftp_host = (
            str(creds.get("sftp_host") or "").strip()
            or getattr(settings, "THE_WHEEL_GROUP_RELAY_HOST", None)
            or DEFAULT_RELAY_HOST
        )
        self.sftp_port = int(
            creds.get("sftp_port")
            or getattr(settings, "THE_WHEEL_GROUP_RELAY_PORT", None)
            or DEFAULT_RELAY_PORT
        )
        self.sftp_user = str(creds.get("ftp_user") or creds.get("sftp_user") or "").strip()
        self.sftp_password = str(
            creds.get("ftp_password") or creds.get("sftp_password") or ""
        ).strip()
        self.sftp_directory = (
            str(creds.get("sftp_directory") or "").strip()
            or getattr(settings, "THE_WHEEL_GROUP_RELAY_DIRECTORY", None)
            or DEFAULT_RELAY_DIRECTORY
        )
        self.local_file_path = local_file_path
        if force_public_share is None:
            force_public_share = bool(
                getattr(settings, "THE_WHEEL_GROUP_FORCE_PUBLIC_SHARE", True)
            )
        self.force_public_share = force_public_share

    # ----------------------------------------------------------------------------------
    # Transport selection
    # ----------------------------------------------------------------------------------
    def source_mode(self) -> str:
        """
        ``sftp`` when this connection carries relay credentials, else ``public_share``.
        ``THE_WHEEL_GROUP_FORCE_PUBLIC_SHARE`` (default True) overrides to the public share: every
        connected company already has relay credentials stored, but TWG is not delivering to the
        relay yet, so without the override every sync would read an empty folder.
        """
        if self.force_public_share:
            return SOURCE_MODE_PUBLIC_SHARE
        if self.sftp_host and self.sftp_user and self.sftp_password:
            return SOURCE_MODE_SFTP
        return SOURCE_MODE_PUBLIC_SHARE

    # ----------------------------------------------------------------------------------
    # Public share transport
    # ----------------------------------------------------------------------------------
    def _http_get(self, url: str, timeout: int) -> bytes:
        request = urllib.request.Request(url, headers=_REQUEST_HEADERS)
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return response.read()
        except urllib.error.HTTPError as e:
            raise exceptions.TheWheelGroupDownloadError(
                "HTTP {} from The Wheel Group share ({}): {}.".format(e.code, url, e.reason)
            )
        except urllib.error.URLError as e:
            raise exceptions.TheWheelGroupDownloadError(
                "Could not reach The Wheel Group share ({}): {}.".format(url, e.reason)
            )

    def _http_reachable(self, url: str, timeout: int) -> None:
        """Connect-time check: the share answers, without pulling the whole ~2 MB archive."""
        request = urllib.request.Request(url, headers=_REQUEST_HEADERS, method="HEAD")
        try:
            urllib.request.urlopen(request, timeout=timeout).close()
        except urllib.error.HTTPError as e:
            raise exceptions.TheWheelGroupDownloadError(
                "HTTP {} from The Wheel Group share ({}): {}.".format(e.code, url, e.reason)
            )
        except urllib.error.URLError as e:
            raise exceptions.TheWheelGroupDownloadError(
                "Could not reach The Wheel Group share ({}): {}.".format(url, e.reason)
            )

    # ----------------------------------------------------------------------------------
    # SFTP (relay) transport
    # ----------------------------------------------------------------------------------
    def _sftp_session(self):
        """Open a paramiko SFTP session. Caller must close the returned (transport, sftp) pair."""
        import paramiko

        try:
            transport = paramiko.Transport((self.sftp_host, self.sftp_port))
            transport.banner_timeout = _SFTP_CONNECT_TIMEOUT_SECONDS
            transport.connect(username=self.sftp_user, password=self.sftp_password)
            sftp = paramiko.SFTPClient.from_transport(transport)
        except Exception as e:
            raise exceptions.TheWheelGroupSFTPConnectionError(
                "Failed to connect to The Wheel Group relay {}:{}: {}.".format(
                    self.sftp_host, self.sftp_port, str(e)
                )
            )
        return transport, sftp

    @staticmethod
    def _close_sftp_session(transport, sftp) -> None:
        for closable in (sftp, transport):
            try:
                if closable is not None:
                    closable.close()
            except Exception as e:  # noqa: BLE001 -- disconnect errors must not mask the real one
                logger.warning("{} Error closing SFTP session: {}.".format(_LOG_PREFIX, str(e)))

    @staticmethod
    def latest_workbook_filename(entries: typing.Iterable[typing.Tuple[str, float]]) -> str:
        """
        Newest mastersheet from ``[(filename, mtime), ...]``. Names matching
        ``WORKBOOK_FILENAME_PATTERN`` win; if the drop is named something else entirely, any
        ``.xlsx``/``.zip`` is accepted rather than failing the sync on a rename.
        """
        candidates = [
            (mtime, name)
            for name, mtime in entries
            if WORKBOOK_FILENAME_PATTERN.search((name or "").strip())
        ]
        if not candidates:
            candidates = [
                (mtime, name)
                for name, mtime in entries
                if (name or "").strip().lower().endswith(_WORKBOOK_SUFFIXES)
            ]
        if not candidates:
            raise exceptions.TheWheelGroupFileNotFoundError(
                "No mastersheet workbook (.xlsx/.zip) found in The Wheel Group feed directory."
            )
        candidates.sort()
        return candidates[-1][1]

    def _relay_entries(self, sftp) -> typing.List[typing.Tuple[str, float]]:
        try:
            return [(attr.filename, attr.st_mtime or 0) for attr in sftp.listdir_attr(self.sftp_directory)]
        except Exception as e:
            raise exceptions.TheWheelGroupDownloadError(
                "Connected to The Wheel Group relay but could not list {!r}: {}.".format(
                    self.sftp_directory, str(e)
                )
            )

    # ----------------------------------------------------------------------------------
    # Download
    # ----------------------------------------------------------------------------------
    @staticmethod
    def _workbook_bytes(filename: str, content: bytes) -> typing.Tuple[str, bytes]:
        """
        Unwrap a download into (workbook filename, xlsx bytes). A zip (what Dropbox serves for a
        folder share) is opened in memory and the mastersheet picked out of it; a bare .xlsx is
        returned as-is.
        """
        if not content:
            raise exceptions.TheWheelGroupDownloadError(
                "The Wheel Group download {} was empty.".format(filename)
            )
        if not content.startswith(b"PK"):
            raise exceptions.TheWheelGroupParseError(
                "The Wheel Group download {} is not a zip or xlsx archive -- the share most "
                "likely returned an HTML page instead of the file.".format(filename)
            )
        try:
            archive = zipfile.ZipFile(io.BytesIO(content))
            names = archive.namelist()
        except zipfile.BadZipFile as e:
            raise exceptions.TheWheelGroupParseError(
                "Could not open The Wheel Group download {}: {}.".format(filename, str(e))
            )

        # An .xlsx is itself a zip, distinguishable by its OOXML workbook part.
        if "xl/workbook.xml" in names:
            archive.close()
            return filename, content

        inner = [
            name
            for name in names
            if name.lower().endswith(".xlsx") and not os.path.basename(name).startswith("._")
        ]
        if not inner:
            archive.close()
            raise exceptions.TheWheelGroupFileNotFoundError(
                "The Wheel Group archive {} contains no .xlsx workbook.".format(filename)
            )
        preferred = [name for name in inner if WORKBOOK_FILENAME_PATTERN.search(name)]
        chosen = sorted(preferred or inner)[0]
        try:
            inner_content = archive.read(chosen)
        finally:
            archive.close()
        return os.path.basename(chosen), inner_content

    def download_latest_workbook(self) -> typing.Tuple[str, bytes]:
        """Download the newest mastersheet over whichever transport this connection uses."""
        if self.local_file_path:
            with open(self.local_file_path, "rb") as f:
                content = f.read()
            return self._workbook_bytes(os.path.basename(self.local_file_path), content)

        mode = self.source_mode()
        if mode == SOURCE_MODE_SFTP:
            transport, sftp = self._sftp_session()
            try:
                filename = self.latest_workbook_filename(self._relay_entries(sftp))
                remote_path = (
                    filename
                    if self.sftp_directory in ("", ".")
                    else "{}/{}".format(self.sftp_directory.rstrip("/"), filename)
                )
                buffer = io.BytesIO()
                sftp.getfo(remote_path, buffer)
                content = buffer.getvalue()
            except exceptions.TheWheelGroupException:
                raise
            except Exception as e:
                raise exceptions.TheWheelGroupDownloadError(
                    "Failed to download The Wheel Group workbook over SFTP: {}.".format(str(e))
                )
            finally:
                self._close_sftp_session(transport, sftp)
        else:
            filename = "the_wheel_group_share.zip"
            content = self._http_get(self.public_share_url, _DOWNLOAD_TIMEOUT_SECONDS)

        workbook_filename, workbook_content = self._workbook_bytes(filename, content)
        logger.info(
            "{} Downloaded {} ({} bytes) via {}.".format(
                _LOG_PREFIX, workbook_filename, len(workbook_content), mode
            )
        )
        return workbook_filename, workbook_content

    def test_connection(self) -> None:
        """
        Connect-time check: confirm the feed source is reachable and (on the relay) actually holds
        a workbook. Never downloads the full archive.
        """
        if self.source_mode() == SOURCE_MODE_SFTP:
            transport, sftp = self._sftp_session()
            try:
                entries = self._relay_entries(sftp)
            finally:
                self._close_sftp_session(transport, sftp)
            self.latest_workbook_filename(entries)
        else:
            if not self.public_share_url:
                raise ValueError("A public share URL is required for The Wheel Group feed.")
            self._http_reachable(self.public_share_url, _HEAD_TIMEOUT_SECONDS)

    # ----------------------------------------------------------------------------------
    # Workbook parsing
    # ----------------------------------------------------------------------------------
    @staticmethod
    def _data_worksheet(workbook):
        for sheet_name in workbook.sheetnames:
            if sheet_name.strip().upper() == DATA_SHEET_NAME.upper():
                return sheet_name, workbook[sheet_name]
        raise exceptions.TheWheelGroupParseError(
            "The Wheel Group workbook has no {!r} worksheet (found: {}).".format(
                DATA_SHEET_NAME, ", ".join(workbook.sheetnames) or "none"
            )
        )

    @staticmethod
    def _header_positions(cells: typing.Sequence[typing.Any]) -> typing.Dict[str, int]:
        """
        Field name -> column index for one header row. The mastersheet repeats ``BRAND`` twice
        (a canonical column and an internal short-code one, "DL" for Dirty Life); ``setdefault``
        keeps the first, which is the canonical name.
        """
        positions: typing.Dict[str, int] = {}
        for index, value in enumerate(cells):
            field = _FIELD_BY_HEADER_KEY.get(_header_key(value))
            if field:
                positions.setdefault(field, index)
        return positions

    @classmethod
    def _parse_data_sheet(cls, worksheet) -> typing.List[typing.Dict]:
        rows: typing.List[typing.Dict] = []
        positions: typing.Optional[typing.Dict[str, int]] = None

        for raw_row in worksheet.iter_rows(values_only=True):
            cells = list(raw_row)
            if positions is None:
                keys = {_header_key(value) for value in cells}
                if all(required in keys for required in _REQUIRED_HEADER_KEYS):
                    positions = cls._header_positions(cells)
                continue

            row = {
                field: _text(cells[index]) if index < len(cells) else None
                for field, index in positions.items()
            }
            if not row.get("sku"):
                continue
            rows.append(row)

        if positions is None:
            raise exceptions.TheWheelGroupParseError(
                "The Wheel Group {!r} worksheet has no SKU/BRAND header row.".format(DATA_SHEET_NAME)
            )
        return rows

    def parse_workbook(self, path: str) -> typing.Dict[str, typing.Any]:
        """Parse a downloaded workbook into ``{"parts": [...], "sheet_name": ...}``."""
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            raise exceptions.TheWheelGroupParseError(
                "Failed to open The Wheel Group workbook {}: {}.".format(path, str(e))
            )

        try:
            sheet_name, worksheet = self._data_worksheet(workbook)
            parts = self._parse_data_sheet(worksheet)
        except exceptions.TheWheelGroupException:
            raise
        except Exception as e:
            raise exceptions.TheWheelGroupParseError(
                "Failed to parse The Wheel Group workbook {}: {}.".format(path, str(e))
            )
        finally:
            workbook.close()

        logger.info(
            "{} Parsed {} rows from {!r}.".format(_LOG_PREFIX, len(parts), sheet_name)
        )
        return {"parts": parts, "sheet_name": sheet_name}

    def get_feed_data(self) -> typing.Dict[str, typing.Any]:
        """
        Download the newest workbook and return its parsed contents plus provenance
        (``source_filename``, ``source_mode``) for the service layer to record on the parts.
        """
        filename, content = self.download_latest_workbook()
        path = os.path.join(
            tempfile.gettempdir(), "the_wheel_group_{}".format(os.path.basename(filename))
        )
        with open(path, "wb") as f:
            f.write(content)
        data = self.parse_workbook(path)
        data["source_filename"] = filename
        data["source_mode"] = "local_file" if self.local_file_path else self.source_mode()
        return data
