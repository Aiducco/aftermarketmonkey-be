"""
Client for Elite Wheel & Tire's inventory workbook.

Elite publishes one ``TriWeeklyUpdate<MM-DD-YYYY>.xlsx`` per drop (three per week) holding the
whole warehouse snapshot: one worksheet per wheel manufacturer (Amani, Azara, Cavallo,
O.E. Revolution, Spec-1, Vortek, XF Dually, XF Off-Road, ...) plus a final ``Tires`` sheet
covering every tire manufacturer. Each sheet carries on-hand counts per Elite location
(Tampa / Atlanta / Miami / Decatur today) and no prices -- the pricing columns only exist on the
per-dealer feed, so ``EliteWheel*CompanyPricing`` stays empty until a dealer account is wired up.

Two transports, same workbook parser:

  * ``public_share`` (default today) -- Elite's SFTPGo web client exposes the same drop folder as
    a public share. The share id lives in ``CompanyProviders.credentials['feed']`` as
    ``public_share_id``/``public_share_url`` (see ``src.constants.ELITE_WHEEL_CREDENTIALS_*``);
    the newest workbook is discovered from the share's JSON directory listing and downloaded over
    plain HTTPS.
  * ``sftp`` -- the real dealer account (``sftp_host``/``sftp_port``/``sftp_user``/
    ``sftp_password``), listing the account home directory and pulling the newest
    ``TriWeeklyUpdate*.xlsx``.

Which one runs is decided per connection by :func:`EliteWheelFeedClient.source_mode`: SFTP as soon
as a connection carries SFTP credentials, otherwise the public share. Flip
``ELITE_WHEEL_FORCE_PUBLIC_SHARE = True`` in settings to pin every connection to the public share
while dealer accounts are still being provisioned.
"""
import datetime
import io
import json
import logging
import os
import re
import tempfile
import typing
import urllib.error
import urllib.parse
import urllib.request

import openpyxl
from django.conf import settings

from src.integrations.clients.elite_wheel import exceptions

logger = logging.getLogger(__name__)

_LOG_PREFIX = "[ELITE-WHEEL-CLIENT]"

DEFAULT_PUBLIC_SHARE_BASE_URL = "https://sftp.elitewheelandtire.com"
DEFAULT_PUBLIC_SHARE_ID = "84ypjmAqgkiZn9T2V4DboM"
DEFAULT_SFTP_HOST = "sftp.elitewheelandtire.com"
DEFAULT_SFTP_PORT = 22

SOURCE_MODE_PUBLIC_SHARE = "public_share"
SOURCE_MODE_SFTP = "sftp"

# Elite's drop filename: TriWeeklyUpdate08-08-2026.xlsx (MM-DD-YYYY). Used to pick the newest
# workbook by its embedded date rather than by string sort, which orders 08-07 after 08-30.
WORKBOOK_FILENAME_PATTERN = re.compile(
    r"^TriWeeklyUpdate(\d{2})-(\d{2})-(\d{4})\.xlsx$", re.IGNORECASE
)

_REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; AfterMarketScout/1.0; +https://aftermarketscout.com)",
    "Accept": "*/*",
}
_DOWNLOAD_TIMEOUT_SECONDS = 300
_LISTING_TIMEOUT_SECONDS = 60

TIRES_SHEET_NAME = "Tires"

# Sheet markers. Every wheel sheet repeats a PART NUMBER header per model/finish block; the Tires
# sheet repeats a Manufacturer header per tire diameter section.
_WHEEL_HEADER_LABEL = "PART NUMBER"
_TIRE_HEADER_LABEL = "MANUFACTURER"
_TIRE_PART_NUMBER_LABEL = "PARTNUMBER"
_BLOCK_TOTAL_LABEL = "TOTAL"

_AS_OF_RE = re.compile(r"as of\s+(\d{1,2}/\d{1,2}/\d{2,4})", re.IGNORECASE)
_MANUFACTURER_RE = re.compile(r"^manufacturer\s*:\s*(.+)$", re.IGNORECASE)
_TIRE_DIAMETER_RE = re.compile(r"tire diameter\s*:\s*([0-9.]+)", re.IGNORECASE)
# Location columns are named "<Location> Available" and the set changes as Elite opens/closes
# warehouses, so they are read off each header row instead of being hardcoded.
_AVAILABLE_COLUMN_RE = re.compile(r"^(.+?)\s+available$", re.IGNORECASE)


def _text(value: typing.Any) -> typing.Optional[str]:
    """Trimmed cell text, or None for blank/whitespace-only cells."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        # openpyxl hands back 66.9 as a float but 20 as 20.0 -- keep whole numbers integral so
        # part numbers and sizes don't pick up a ".0" tail.
        text = str(int(value))
    else:
        text = str(value).strip()
    text = text.strip()
    return text or None


def _int_or_none(value: typing.Any) -> typing.Optional[int]:
    text = _text(value)
    if text is None:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


class EliteWheelFeedClient(object):
    """
    Downloads Elite's newest inventory workbook and returns raw row dicts, one list for wheels and
    one for tires (numeric coercion and model mapping are the service layer's job, matching the
    Quadratec and Rough Country clients).
    """

    def __init__(
        self,
        credentials: typing.Optional[typing.Dict] = None,
        local_file_path: typing.Optional[str] = None,
        force_public_share: typing.Optional[bool] = None,
    ):
        creds = credentials or {}
        self.public_share_base_url = (
            str(creds.get("public_share_url") or "").strip()
            or getattr(settings, "ELITE_WHEEL_PUBLIC_SHARE_URL", None)
            or DEFAULT_PUBLIC_SHARE_BASE_URL
        ).rstrip("/")
        self.public_share_id = (
            str(creds.get("public_share_id") or "").strip()
            or getattr(settings, "ELITE_WHEEL_PUBLIC_SHARE_ID", None)
            or DEFAULT_PUBLIC_SHARE_ID
        )
        self.sftp_host = str(creds.get("sftp_host") or "").strip()
        self.sftp_port = int(creds.get("sftp_port") or DEFAULT_SFTP_PORT)
        self.sftp_user = str(creds.get("sftp_user") or "").strip()
        self.sftp_password = str(creds.get("sftp_password") or "").strip()
        self.sftp_directory = str(creds.get("sftp_directory") or "").strip() or "."
        self.local_file_path = local_file_path
        if force_public_share is None:
            force_public_share = bool(getattr(settings, "ELITE_WHEEL_FORCE_PUBLIC_SHARE", False))
        self.force_public_share = force_public_share

    # ----------------------------------------------------------------------------------
    # Transport selection
    # ----------------------------------------------------------------------------------
    def source_mode(self) -> str:
        """
        ``sftp`` when this connection carries a full set of SFTP credentials, else
        ``public_share``. ``ELITE_WHEEL_FORCE_PUBLIC_SHARE`` overrides to the public share so a
        dealer's credentials can be stored and validated before the SFTP ingest is switched on.
        """
        if self.force_public_share:
            return SOURCE_MODE_PUBLIC_SHARE
        if self.sftp_host and self.sftp_user and self.sftp_password:
            return SOURCE_MODE_SFTP
        return SOURCE_MODE_PUBLIC_SHARE

    # ----------------------------------------------------------------------------------
    # Public share transport
    # ----------------------------------------------------------------------------------
    def _public_share_path(self, suffix: str) -> str:
        return "{}/web/client/pubshares/{}/{}".format(
            self.public_share_base_url, self.public_share_id, suffix.lstrip("/")
        )

    def _public_share_listing_url(self) -> str:
        return self._public_share_path("dirs?path=%2F")

    def _public_share_download_url(self, filename: str) -> str:
        return self._public_share_path(
            "browse?path={}".format(urllib.parse.quote("/" + filename, safe=""))
        )

    def _http_get(self, url: str, timeout: int) -> bytes:
        request = urllib.request.Request(url, headers=_REQUEST_HEADERS)
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return response.read()
        except urllib.error.HTTPError as e:
            raise exceptions.EliteWheelDownloadError(
                "HTTP {} from Elite public share ({}): {}.".format(e.code, url, e.reason)
            )
        except urllib.error.URLError as e:
            raise exceptions.EliteWheelDownloadError(
                "Could not reach Elite public share ({}): {}.".format(url, e.reason)
            )

    def _public_share_filenames(self) -> typing.List[str]:
        raw = self._http_get(self._public_share_listing_url(), _LISTING_TIMEOUT_SECONDS)
        try:
            entries = json.loads(raw.decode("utf-8"))
        except (UnicodeDecodeError, ValueError) as e:
            raise exceptions.EliteWheelParseError(
                "Elite public share directory listing was not valid JSON: {}.".format(str(e))
            )
        if not isinstance(entries, list):
            raise exceptions.EliteWheelParseError(
                "Elite public share directory listing was not a list of entries."
            )
        return [
            str(entry.get("name") or "")
            for entry in entries
            if isinstance(entry, dict) and entry.get("name")
        ]

    # ----------------------------------------------------------------------------------
    # SFTP transport
    # ----------------------------------------------------------------------------------
    def _sftp_session(self):
        """Open a paramiko SFTP session. Caller must close the returned (transport, sftp) pair."""
        import paramiko

        try:
            transport = paramiko.Transport((self.sftp_host, self.sftp_port))
            transport.connect(username=self.sftp_user, password=self.sftp_password)
            sftp = paramiko.SFTPClient.from_transport(transport)
        except Exception as e:
            raise exceptions.EliteWheelSFTPConnectionError(
                "Failed to connect to Elite SFTP {}:{}: {}.".format(
                    self.sftp_host, self.sftp_port, str(e)
                )
            )
        return transport, sftp

    @staticmethod
    def _close_sftp_session(transport, sftp) -> None:
        for closable in (sftp, transport):
            try:
                if closable is not None:
                    closable.close()
            except Exception as e:  # noqa: BLE001 -- disconnect errors must not mask the real one
                logger.warning("{} Error closing SFTP session: {}.".format(_LOG_PREFIX, str(e)))

    # ----------------------------------------------------------------------------------
    # Workbook discovery
    # ----------------------------------------------------------------------------------
    @staticmethod
    def latest_workbook_filename(filenames: typing.Iterable[str]) -> str:
        """
        Newest ``TriWeeklyUpdate<MM-DD-YYYY>.xlsx`` by its embedded date. Elite keeps roughly a
        month of drops in the folder, and their names sort wrong lexicographically across a month
        boundary, so the date is parsed rather than the string compared.
        """
        candidates: typing.List[typing.Tuple[datetime.date, str]] = []
        for filename in filenames:
            match = WORKBOOK_FILENAME_PATTERN.match((filename or "").strip())
            if not match:
                continue
            month, day, year = (int(group) for group in match.groups())
            try:
                candidates.append((datetime.date(year, month, day), filename))
            except ValueError:
                logger.warning(
                    "{} Skipping workbook with impossible date in name: {}.".format(
                        _LOG_PREFIX, filename
                    )
                )
        if not candidates:
            raise exceptions.EliteWheelFileNotFoundError(
                "No TriWeeklyUpdate*.xlsx workbook found in the Elite feed directory."
            )
        candidates.sort()
        return candidates[-1][1]

    def download_latest_workbook(self) -> typing.Tuple[str, bytes]:
        """Download the newest workbook over whichever transport this connection uses."""
        if self.local_file_path:
            with open(self.local_file_path, "rb") as f:
                content = f.read()
            return os.path.basename(self.local_file_path), content

        mode = self.source_mode()
        if mode == SOURCE_MODE_SFTP:
            transport, sftp = self._sftp_session()
            try:
                filename = self.latest_workbook_filename(sftp.listdir(self.sftp_directory))
                remote_path = (
                    filename
                    if self.sftp_directory == "."
                    else "{}/{}".format(self.sftp_directory.rstrip("/"), filename)
                )
                buffer = io.BytesIO()
                sftp.getfo(remote_path, buffer)
                content = buffer.getvalue()
            except exceptions.EliteWheelException:
                raise
            except Exception as e:
                raise exceptions.EliteWheelDownloadError(
                    "Failed to download Elite workbook over SFTP: {}.".format(str(e))
                )
            finally:
                self._close_sftp_session(transport, sftp)
        else:
            filename = self.latest_workbook_filename(self._public_share_filenames())
            content = self._http_get(
                self._public_share_download_url(filename), _DOWNLOAD_TIMEOUT_SECONDS
            )

        if not content:
            raise exceptions.EliteWheelDownloadError(
                "Elite workbook {} downloaded empty.".format(filename)
            )
        logger.info(
            "{} Downloaded {} ({} bytes) via {}.".format(_LOG_PREFIX, filename, len(content), mode)
        )
        return filename, content

    def test_connection(self) -> None:
        """
        Connect-time check: confirm the feed directory is reachable and holds at least one
        recognisable workbook. Never downloads the full ~4 MB file.
        """
        if self.source_mode() == SOURCE_MODE_SFTP:
            transport, sftp = self._sftp_session()
            try:
                filenames = sftp.listdir(self.sftp_directory)
            except Exception as e:
                raise exceptions.EliteWheelDownloadError(
                    "Connected to Elite SFTP but could not list {!r}: {}.".format(
                        self.sftp_directory, str(e)
                    )
                )
            finally:
                self._close_sftp_session(transport, sftp)
        else:
            if not self.public_share_id:
                raise ValueError("public_share_id is required for the Elite public share feed.")
            filenames = self._public_share_filenames()
        self.latest_workbook_filename(filenames)

    # ----------------------------------------------------------------------------------
    # Workbook parsing
    # ----------------------------------------------------------------------------------
    @staticmethod
    def _header_index(cells: typing.Sequence[typing.Any]) -> typing.Dict[str, int]:
        """
        Uppercased header label -> first column index. Wheel sheets repeat ``BP`` twice (inner and
        outer bolt pattern); only the first is recorded here and the second is read from
        :func:`_repeated_header_indexes`.
        """
        index: typing.Dict[str, int] = {}
        for position, value in enumerate(cells):
            label = _text(value)
            if label:
                index.setdefault(label.upper(), position)
        return index

    @staticmethod
    def _repeated_header_indexes(
        cells: typing.Sequence[typing.Any], label: str
    ) -> typing.List[int]:
        wanted = label.upper()
        return [
            position
            for position, value in enumerate(cells)
            if (_text(value) or "").upper() == wanted
        ]

    @staticmethod
    def _location_columns(
        cells: typing.Sequence[typing.Any],
    ) -> typing.List[typing.Tuple[str, int]]:
        """``[(location label, column index), ...]`` for every ``<Location> Available`` column."""
        locations: typing.List[typing.Tuple[str, int]] = []
        for position, value in enumerate(cells):
            label = _text(value)
            if not label:
                continue
            match = _AVAILABLE_COLUMN_RE.match(label)
            if match:
                locations.append((match.group(1).strip(), position))
        return locations

    @staticmethod
    def _cell(cells: typing.Sequence[typing.Any], position: typing.Optional[int]) -> typing.Any:
        if position is None or position >= len(cells):
            return None
        return cells[position]

    @classmethod
    def _availability(
        cls,
        cells: typing.Sequence[typing.Any],
        locations: typing.Sequence[typing.Tuple[str, int]],
    ) -> typing.Dict[str, int]:
        """On-hand per location for one data row; locations with no number are reported as 0."""
        return {
            label: (_int_or_none(cls._cell(cells, position)) or 0)
            for label, position in locations
        }

    def _parse_wheel_sheet(self, worksheet) -> typing.List[typing.Dict]:
        """
        One wheel manufacturer's sheet: a ``Manufacturer: <name>`` banner, then repeating blocks of
        [model/finish title row, PART NUMBER header row, data rows, TOTAL row].
        """
        rows: typing.List[typing.Dict] = []
        manufacturer: typing.Optional[str] = None
        group_label: typing.Optional[str] = None
        header: typing.Optional[typing.Dict[str, int]] = None
        locations: typing.List[typing.Tuple[str, int]] = []
        bolt_pattern_positions: typing.List[int] = []

        for raw_row in worksheet.iter_rows(values_only=True):
            cells = list(raw_row)
            populated = [(i, _text(v)) for i, v in enumerate(cells) if _text(v)]
            if not populated:
                header = None
                continue

            first_text = populated[0][1]
            manufacturer_match = _MANUFACTURER_RE.match(first_text)
            if manufacturer_match:
                manufacturer = manufacturer_match.group(1).strip()
                header = None
                continue

            if first_text.upper() == _WHEEL_HEADER_LABEL:
                header = self._header_index(cells)
                locations = self._location_columns(cells)
                bolt_pattern_positions = self._repeated_header_indexes(cells, "BP")
                continue

            if header is None:
                # Between blocks: a lone cell is the next model/finish title (e.g.
                # "XF-211 Gloss Black & Milled"); banners and spacer rows are ignored.
                if len(populated) == 1:
                    group_label = first_text
                continue

            part_number = _text(self._cell(cells, header.get(_WHEEL_HEADER_LABEL)))
            if not part_number or part_number.upper() == _BLOCK_TOTAL_LABEL:
                header = None
                continue

            bolt_patterns = [
                _text(self._cell(cells, position)) for position in bolt_pattern_positions
            ]
            rows.append(
                {
                    "manufacturer": manufacturer,
                    "group_label": group_label,
                    "part_number": part_number,
                    "size": _text(self._cell(cells, header.get("SIZE"))),
                    "bolt_pattern_1": bolt_patterns[0] if len(bolt_patterns) > 0 else None,
                    "bolt_pattern_2": bolt_patterns[1] if len(bolt_patterns) > 1 else None,
                    "offset": _text(self._cell(cells, header.get("ET"))),
                    "center_bore": _text(self._cell(cells, header.get("CB"))),
                    "finish": _text(self._cell(cells, header.get("FINISH"))),
                    "availability": self._availability(cells, locations),
                }
            )
        return rows

    def _parse_tire_sheet(self, worksheet) -> typing.List[typing.Dict]:
        """
        The ``Tires`` sheet: repeating ``Inventory for Tire Diameter: <n>`` sections, each with a
        Manufacturer / Model / PartNumber / Raw Size header and its data rows.
        """
        rows: typing.List[typing.Dict] = []
        tire_diameter: typing.Optional[str] = None
        header: typing.Optional[typing.Dict[str, int]] = None
        locations: typing.List[typing.Tuple[str, int]] = []

        for raw_row in worksheet.iter_rows(values_only=True):
            cells = list(raw_row)
            populated = [(i, _text(v)) for i, v in enumerate(cells) if _text(v)]
            if not populated:
                header = None
                continue

            first_text = populated[0][1]
            diameter_match = _TIRE_DIAMETER_RE.search(first_text)
            if diameter_match:
                tire_diameter = diameter_match.group(1)
                header = None
                continue

            index = self._header_index(cells)
            if _TIRE_HEADER_LABEL in index and _TIRE_PART_NUMBER_LABEL in index:
                header = index
                locations = self._location_columns(cells)
                continue

            if header is None:
                continue

            part_number = _text(self._cell(cells, header.get(_TIRE_PART_NUMBER_LABEL)))
            manufacturer = _text(self._cell(cells, header.get(_TIRE_HEADER_LABEL)))
            if not part_number or not manufacturer:
                header = None
                continue

            rows.append(
                {
                    "manufacturer": manufacturer,
                    "model": _text(self._cell(cells, header.get("MODEL"))),
                    "part_number": part_number,
                    "raw_size": _text(self._cell(cells, header.get("RAW SIZE"))),
                    "tire_diameter": tire_diameter,
                    "availability": self._availability(cells, locations),
                }
            )
        return rows

    @staticmethod
    def _as_of_date(worksheet) -> typing.Optional[datetime.date]:
        """The ``As of MM/DD/YYYY`` banner near the top of every sheet."""
        for raw_row in worksheet.iter_rows(min_row=1, max_row=20, values_only=True):
            for value in raw_row:
                text = _text(value)
                if not text:
                    continue
                match = _AS_OF_RE.search(text)
                if not match:
                    continue
                for fmt in ("%m/%d/%Y", "%m/%d/%y"):
                    try:
                        return datetime.datetime.strptime(match.group(1), fmt).date()
                    except ValueError:
                        continue
        return None

    def parse_workbook(self, path: str) -> typing.Dict[str, typing.Any]:
        """
        Parse a downloaded workbook into ``{"wheels": [...], "tires": [...], "as_of": date|None,
        "locations": [...]}``. Sheets are classified by name: everything except ``Tires`` is a
        wheel manufacturer.
        """
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            raise exceptions.EliteWheelParseError(
                "Failed to open Elite workbook {}: {}.".format(path, str(e))
            )

        wheels: typing.List[typing.Dict] = []
        tires: typing.List[typing.Dict] = []
        as_of: typing.Optional[datetime.date] = None
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                if as_of is None:
                    as_of = self._as_of_date(worksheet)
                    worksheet = workbook[sheet_name]  # read-only sheets are single-pass
                if sheet_name.strip().upper() == TIRES_SHEET_NAME.upper():
                    tires.extend(self._parse_tire_sheet(worksheet))
                else:
                    for row in self._parse_wheel_sheet(worksheet):
                        # The per-sheet banner is the authoritative manufacturer name; the tab
                        # title is the fallback for a sheet that ever ships without one.
                        if not row.get("manufacturer"):
                            row["manufacturer"] = sheet_name.strip()
                        row["sheet_name"] = sheet_name.strip()
                        wheels.append(row)
        except exceptions.EliteWheelException:
            raise
        except Exception as e:
            raise exceptions.EliteWheelParseError(
                "Failed to parse Elite workbook {}: {}.".format(path, str(e))
            )
        finally:
            workbook.close()

        locations = sorted(
            {label for row in wheels + tires for label in (row.get("availability") or {})}
        )
        logger.info(
            "{} Parsed {} wheel rows and {} tire rows (as_of={}, locations={}).".format(
                _LOG_PREFIX, len(wheels), len(tires), as_of, ", ".join(locations) or "none"
            )
        )
        return {"wheels": wheels, "tires": tires, "as_of": as_of, "locations": locations}

    def get_feed_data(self) -> typing.Dict[str, typing.Any]:
        """
        Download the newest workbook and return its parsed contents plus provenance
        (``source_filename``, ``source_mode``) for the service layer to record on the raw rows.
        """
        filename, content = self.download_latest_workbook()
        path = os.path.join(
            tempfile.gettempdir(), "elite_wheel_{}".format(os.path.basename(filename))
        )
        with open(path, "wb") as f:
            f.write(content)
        data = self.parse_workbook(path)
        data["source_filename"] = filename
        data["source_mode"] = "local_file" if self.local_file_path else self.source_mode()
        return data
