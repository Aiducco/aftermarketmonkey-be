"""
Read a file the brand sent us: CSV, TSV or XLSX.

This is the department's workhorse, because "how do we get your tire data" is answered with a
spreadsheet far more often than with an API. Everything about a given file lives in the registry
row's ``config`` -- so onboarding a brand that emails a sheet is a row and a column map, with no
code and no deploy.

Config::

    {
      "path": "resources/brand_data/michelin/2026-q3-us.xlsx",   # or "paths": [...]
      "sheet": "Passenger",          # XLSX only; default is the first sheet
      "header_row": 1,               # 1-based; rows above it are preamble and are skipped
      "encoding": "utf-8-sig",       # text files only; the -sig strips Excel's BOM
      "delimiter": ",",              # default: tab for .tsv, comma otherwise
      "row_filter": {"Type": ["Tire", "TIRE"]},
      "field_map": { ... }
    }

Paths are resolved against the repo root when relative. Files belong in ``resources/brand_data/``
and dated: a manufacturer sheet is a snapshot, the next one supersedes it rather than amending it,
and knowing which edition a row came from is the difference between a stale spec and a wrong one.

``row_filter`` matters more than it looks. Brand sheets routinely carry the whole product line --
tires, wheels, accessories, discontinued rows, a totals row at the bottom -- and the filter is
what keeps those out of a tire table. Matching is loose on case and spacing, like the column map.
"""
import csv
import hashlib
import pathlib
import typing

from django.conf import settings

from src.integrations.brand_data import base, normalize
from src.integrations.brand_data.registry import loader

TEXT_SUFFIXES = frozenset([".csv", ".tsv", ".txt"])
EXCEL_SUFFIXES = frozenset([".xlsx", ".xlsm"])


@loader("csv")
def load(ctx: base.LoaderContext) -> typing.Iterator[base.SourceRecord]:
    paths = _paths(ctx)
    ctx.input_label = ", ".join(str(path) for path in paths)
    produced = 0
    for path in paths:
        for record in _read(ctx, path):
            yield record
            produced += 1
            if ctx.limit is not None and produced >= ctx.limit:
                return


def fingerprint(ctx: base.LoaderContext) -> str:
    """
    The digest of the file(s) without parsing them, so an unchanged quarterly sheet costs a read
    and nothing else. Built the same way ``LoaderContext.observe`` builds the run's fingerprint --
    sha256 over the bytes, in path order -- because ``ingest`` compares the two directly.
    """
    digest = hashlib.sha256()
    for path in _paths(ctx):
        digest.update(path.read_bytes())
    return digest.hexdigest()


load.fingerprint = fingerprint


def _paths(ctx: base.LoaderContext) -> typing.List[pathlib.Path]:
    if ctx.file_override:
        raw_paths: typing.Sequence[str] = [ctx.file_override]
    else:
        raw_paths = ctx.config.get("paths") or ([ctx.config["path"]] if ctx.config.get("path") else [])
    if not raw_paths:
        raise base.SourceConfigError(f"{ctx.source.slug}: config needs 'path' (or 'paths'), or a --file override")

    resolved = []
    for raw in raw_paths:
        path = pathlib.Path(raw)
        if not path.is_absolute():
            path = pathlib.Path(settings.BASE_DIR) / path
        if not path.exists():
            raise base.SourceFetchError(
                f"{ctx.source.slug}: {path} does not exist. Manufacturer files are snapshots that "
                f"arrive by hand -- check resources/brand_data/ for the edition that did."
            )
        resolved.append(path)
    return resolved


def _read(ctx: base.LoaderContext, path: pathlib.Path) -> typing.Iterator[base.SourceRecord]:
    suffix = path.suffix.lower()
    if suffix in EXCEL_SUFFIXES:
        rows = _excel_rows(ctx, path)
    elif suffix in TEXT_SUFFIXES:
        rows = _text_rows(ctx, path)
    else:
        raise base.SourceConfigError(
            f"{ctx.source.slug}: {path.name} is a {suffix or 'suffixless'} file; this loader reads "
            f"{', '.join(sorted(TEXT_SUFFIXES | EXCEL_SUFFIXES))}"
        )

    header_row = int(ctx.config.get("header_row") or 1)
    row_filter = _compile_filter(ctx.config.get("row_filter") or {})
    headers: typing.List[str] = []

    for line_number, values in enumerate(rows, start=1):
        if line_number < header_row:
            continue
        if line_number == header_row:
            headers = [normalize.clean(value) or f"column_{index + 1}" for index, value in enumerate(values)]
            continue
        payload = {header: value for header, value in zip(headers, values) if normalize.clean(value) is not None}
        if not payload:
            continue  # a spacer row, of which brand sheets have many
        if not row_filter(payload):
            continue
        yield base.SourceRecord(payload=payload, label=f"{path.name}:{line_number}")


def _text_rows(ctx: base.LoaderContext, path: pathlib.Path) -> typing.Iterator[typing.Sequence[str]]:
    ctx.observe(path.read_bytes())
    encoding = ctx.config.get("encoding") or "utf-8-sig"
    delimiter = ctx.config.get("delimiter") or ("\t" if path.suffix.lower() == ".tsv" else ",")
    with path.open("r", encoding=encoding, newline="") as handle:
        yield from csv.reader(handle, delimiter=delimiter)


def _excel_rows(ctx: base.LoaderContext, path: pathlib.Path) -> typing.Iterator[typing.Sequence[str]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - openpyxl is a pinned dependency
        raise base.SourceConfigError("openpyxl is required to read .xlsx sources") from exc

    ctx.observe(path.read_bytes())
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = ctx.config.get("sheet")
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise base.SourceConfigError(
                f"{ctx.source.slug}: {path.name} has no sheet {sheet_name!r} (it has: {', '.join(workbook.sheetnames)})"
            )
        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        for row in sheet.iter_rows(values_only=True):
            yield ["" if cell is None else str(cell) for cell in row]
    finally:
        workbook.close()


def _compile_filter(spec: typing.Mapping[str, typing.Any]) -> typing.Callable[[typing.Dict[str, str]], bool]:
    """
    ``{"Type": ["Tire"]}`` -> keep rows whose Type is Tire. Column and value both matched loosely.

    A row missing the filtered column entirely is dropped, not kept: the column is how the sheet
    says what the row is, and a row that does not say is not one we can claim is a tire.
    """
    if not spec:
        return lambda payload: True

    wanted = {
        _loose(column): {_loose(value) for value in (values if isinstance(values, (list, tuple)) else [values])}
        for column, values in spec.items()
    }

    def matches(payload: typing.Dict[str, str]) -> bool:
        loose_payload = {_loose(key): _loose(value) for key, value in payload.items()}
        return all(loose_payload.get(column) in values for column, values in wanted.items())

    return matches


def _loose(value: typing.Any) -> str:
    return "".join(character for character in str(value).lower() if character.isalnum())
